"""Create the Excel template used to fill in the invoices."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

# File written in the current folder.
TEMPLATE = "faturas.xlsx"

# Template columns (Portuguese, as used by the business). Only client and value for now.
COLUMNS = ["cliente", "valor"]


def main() -> None:
    """Write faturas.xlsx with the header row and one example row."""
    target = Path(TEMPLATE)
    if target.exists():
        print(f"{target} already exists, nothing changed.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "faturas"

    # Bold header.
    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Example row showing the expected format.
    ws.append(["Cliente Exemplo", 100.00])
    ws["B2"].number_format = "0.00"

    # Column widths, only for readability.
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 12

    wb.save(target)
    print(f"Template created: {target}")


if __name__ == "__main__":
    main()
