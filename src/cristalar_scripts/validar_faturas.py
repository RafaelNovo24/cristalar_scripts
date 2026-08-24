"""Fill the PC Plus invoice form from an Excel file.

Every client is filled in its own tab, all tabs working at the same time in a
single browser window. Dry-run by default: the form is filled and left open,
Finalizar is only pressed when SUBMIT is True.
Usage: python -m cristalar_scripts.validar_faturas [faturas.xlsx]
"""

import asyncio
import os
import sys
from getpass import getpass
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import load_workbook
from playwright.async_api import async_playwright

from cristalar_scripts import paths
from cristalar_scripts.browser_setup import ensure_chromium

# ---------------------------------------------------------------------------
# SELECTORS - the HTML elements of the site, all in one place.
# ---------------------------------------------------------------------------
SELECTORS = {
    # Login page (/Account/Login).
    "login_email_input": "#Email",
    "login_password_input": "#Password",
    "login_entrar_button": "input[type='submit'][value='Entrar']",
    # Home page toolbar.
    "nova_fatura_button": "a[href='/Invoices/Create']",
    # "Tipo de fatura" radio group. The real radio is invisible (iCheck draws
    # its own control on top), so we click the helper overlay instead.
    "tipo_fatura_radio": "#fatura + ins.iCheck-helper",
    # select2 comboboxes: "Cliente" and, in the Items section, "Artigo".
    "cliente_dropdown": "#select2-cliente-container",
    "artigo_dropdown": "#select2-produto-container",
    "dropdown_search_input": "input.select2-search__field",
    # Suggestions of the open dropdown. The "Sem resultados" line is also an
    # <li class="select2-results__option">, but it carries no aria-selected,
    # so requiring that attribute keeps only the real, clickable suggestions.
    "dropdown_options": "li.select2-results__option[aria-selected]",
    # Items section: "Adicionar" creates the item line, which then shows the
    # "Preço Unitário" box. Its id carries a random GUID, so we go by class.
    "adicionar_item_button": "#add-line",
    "unit_price_input": "div.product_total input.product_price",
    # Toolbar button that closes and issues the invoice.
    "finalizar_button": "button[value='Finalizar']",
}

# Item line used in every invoice. The dropdown also holds
# "Prestacao de Servico de Limpeza- ES", so the match must be exact.
ARTIGO = "Prestação de Serviço de Limpeza"

# Site address - always the same, so it lives here and not in .env.
HOME_URL = "http://www.pcplusonline.pt/Home/Index"

# Without a valid session the site redirects to this path.
LOGIN_PATH = "/Account/Login"

# Excel file used when no argument is given.
DEFAULT_EXCEL = "faturas.xlsx"

# Visible browser so we can follow the filling.
HEADLESS = False

# How many client tabs work at the same time. Keep it low enough for the site
# (and the machine) to cope; the remaining rows wait for a free slot.
PARALLEL_TABS = 5

# How long to wait for a dropdown to filter its suggestions (milliseconds).
DROPDOWN_WAIT_MS = 2000

# How long to wait for the item line to appear after "Adicionar" (ms).
ITEM_LINE_WAIT_MS = 2000

# Press "Finalizar" at the end. False keeps the run a dry-run: the invoice is
# filled and left open, nothing is issued. Set to True to really submit.
SUBMIT = False


def read_invoices(path):
    """Read the Excel file and return a list of (client, value) tuples."""
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    invoices = []
    # min_row=2 skips the header row (cliente / valor).
    for client, value in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        if client is None and value is None:
            continue  # empty row
        invoices.append((str(client).strip(), float(value)))
    return invoices


def get_credentials():
    """Return (user, password) from .env, asking the user for what is missing."""
    user = os.getenv("CRISTALAR_USER") or input("User: ")
    password = os.getenv("CRISTALAR_PASSWORD") or getpass("Password: ")
    return user, password


def is_on_login_page(page):
    """True when the site redirected us to the login page."""
    # Going to HOME_URL without a valid session lands on
    # /Account/Login?ReturnUrl=%2fHome%2fIndex, so the URL is enough to tell.
    return LOGIN_PATH.lower() in page.url.lower()


async def submit_login_form(page, credentials):
    """Fill Email/Password, press Entrar and save the session to STATE_FILE."""
    user, password = credentials
    await page.fill(SELECTORS["login_email_input"], user)
    await page.fill(SELECTORS["login_password_input"], password)
    await page.click(SELECTORS["login_entrar_button"])
    await page.wait_for_load_state("networkidle")
    # The login sends us back to ReturnUrl (the home page); still being on the
    # login page means the credentials were refused.
    if is_on_login_page(page):
        raise RuntimeError("Login failed: still on the login page.")
    paths.ensure_data_dir()
    await page.context.storage_state(path=paths.STATE_FILE)
    print("Logged in, session saved.")


async def click_nova_fatura_button(page, steps):
    """Go to the home page and press Nova Fatura to open an empty form."""
    await page.goto(HOME_URL)
    await page.click(SELECTORS["nova_fatura_button"])
    await page.wait_for_load_state("networkidle")
    steps.append("opened the Nova Fatura form")


async def check_tipo_fatura_radio(page, steps):
    """Tick the Fatura option of the Tipo de fatura radio group."""
    await page.click(SELECTORS["tipo_fatura_radio"])
    steps.append("ticked Tipo de fatura = Fatura")


async def search_in_dropdown(page, dropdown, text):
    """Open a select2 dropdown, type text and return the suggestions locator.

    The text is typed key by key: select2 only filters on real keyboard
    events, so setting the value directly would leave the full list on screen.
    """
    await page.click(dropdown)
    await page.keyboard.type(text, delay=30)
    await page.wait_for_timeout(DROPDOWN_WAIT_MS)
    return page.locator(SELECTORS["dropdown_options"])


async def select_cliente(page, client, steps):
    """Pick the client in the Cliente dropdown by typing its name.

    Returns True when exactly one suggestion matched and was selected.
    Returns False (with a warning) on zero matches (client does not exist) or
    on several matches (the name is too vague to know which one is meant).
    """
    options = await search_in_dropdown(page, SELECTORS["cliente_dropdown"], client)
    count = await options.count()

    if count == 0:
        steps.append(f"WARNING: no client found for '{client}', skipped")
        return False
    if count > 1:
        names = [await options.nth(i).inner_text() for i in range(min(count, 5))]
        steps.append(f"WARNING: {count} clients match '{client}' "
                     f"({', '.join(names)} ...), name is not specific "
                     "enough, skipped")
        return False

    await options.first.click()
    steps.append(f"selected Cliente = {client}")
    return True


async def select_artigo(page, steps):
    """Pick the fixed item (ARTIGO) in the Artigo dropdown of the Items table."""
    options = await search_in_dropdown(page, SELECTORS["artigo_dropdown"], ARTIGO)
    # Typing the full name also matches "... - ES", so take the suggestion
    # whose text is exactly ARTIGO.
    for index in range(await options.count()):
        option = options.nth(index)
        if (await option.inner_text()).strip() == ARTIGO:
            await option.click()
            steps.append(f"selected Artigo = {ARTIGO}")
            return
    raise RuntimeError(f"item '{ARTIGO}' not found in the Artigo dropdown")


async def click_adicionar_item_button(page, steps):
    """Press "Adicionar" to create the item line with the price boxes."""
    await page.click(SELECTORS["adicionar_item_button"])
    # The line is added by JavaScript, so wait for the price box to show up.
    await page.wait_for_selector(SELECTORS["unit_price_input"],
                                 timeout=ITEM_LINE_WAIT_MS + 5000)
    await page.wait_for_timeout(ITEM_LINE_WAIT_MS)
    steps.append("clicked Adicionar, item line created")


async def fill_preco_unitario(page, value, steps):
    """Write the amount in the "Preço Unitário" box of the last item line."""
    # The site uses the Portuguese decimal separator (the box starts at
    # "0,0000"), so send the value with a comma.
    amount = f"{value:.2f}".replace(".", ",")
    price_box = page.locator(SELECTORS["unit_price_input"]).last
    await price_box.fill(amount)
    # The line total is only recalculated when the box loses focus, so leave
    # it with Tab instead of jumping straight to the next step.
    await price_box.press("Tab")
    await page.wait_for_timeout(500)
    steps.append(f"filled Preço Unitário = {amount}")


async def click_finalizar_button(page, steps):
    """Press "Finalizar" to close and issue the invoice."""
    await page.click(SELECTORS["finalizar_button"])
    await page.wait_for_load_state("networkidle")
    steps.append("clicked Finalizar, invoice submitted")


async def fill_invoice(context, client, value, index, slots, submit, on_event=None):
    """Fill one invoice in its own tab. True when filled, False when skipped.

    The steps of this invoice are collected in a list and reported together at
    the end, so the tabs running side by side do not mix up their output.
    """
    steps = []
    result = False
    # slots limits how many tabs work at the same time.
    async with slots:
        page = await context.new_page()
        try:
            await click_nova_fatura_button(page, steps)
            await check_tipo_fatura_radio(page, steps)

            if await select_cliente(page, client, steps):
                await select_artigo(page, steps)
                await click_adicionar_item_button(page, steps)
                await fill_preco_unitario(page, value, steps)

                # The form is left open at this point, whether or not it gets
                # submitted, so it can be reviewed in its own tab.
                if submit:
                    await click_finalizar_button(page, steps)
                else:
                    steps.append("dry-run: Finalizar not pressed")
                result = True
            else:
                await page.close()  # nothing to review in this tab
        except Exception as error:  # noqa: BLE001 - keep the other tabs going
            steps.append(f"FAILED: {error}")
            result = None

    if on_event is not None:
        on_event({
            "index": index,
            "client": client,
            "value": value,
            "steps": steps,
            "result": result,
        })
    return result


def print_event(event):
    """Report one fill_invoice event in the CLI's step-by-step format."""
    print(f"[{event['index']}] {event['client']} - {event['value']:.2f}")
    for step in event["steps"]:
        print(f"    - {step}")


class InvoiceSession:
    """A logged-in browser session that can fill invoices across several calls.

    Splitting start/fill/close (instead of one function wrapped in a single
    "with" block) lets the browser stay open after the invoices are filled,
    so a caller such as the Streamlit UI can leave it up for review and close
    it later, from a different call, on its own event loop thread.
    """

    def __init__(self):
        self._playwright = None
        self._browser = None
        self._context = None

    async def start(self, credentials, headless=HEADLESS, on_event=None):
        """Launch the browser and make sure we are logged in."""
        on_line = (lambda line: on_event({"log": line})) if on_event else None

        self._playwright = await async_playwright().start()
        await ensure_chromium(self._playwright, on_line)
        self._browser = await self._playwright.chromium.launch(headless=headless)

        # Reuse the saved session when we have one.
        if paths.STATE_FILE.exists():
            self._context = await self._browser.new_context(
                storage_state=paths.STATE_FILE)
        else:
            self._context = await self._browser.new_context()

        login_page = await self._context.new_page()
        await login_page.goto(HOME_URL)
        if is_on_login_page(login_page):
            await submit_login_form(login_page, credentials)
        elif on_line is not None:
            on_line("Already logged in.")
        await login_page.close()

    async def fill(self, invoices, submit=SUBMIT, parallel_tabs=PARALLEL_TABS,
                   on_event=None):
        """Fill every invoice in its own tab, all at once. Returns the results."""
        slots = asyncio.Semaphore(parallel_tabs)
        return await asyncio.gather(*[
            fill_invoice(self._context, client, value, index, slots, submit, on_event)
            for index, (client, value) in enumerate(invoices, start=1)
        ])

    async def close(self):
        """Close the browser and stop Playwright."""
        await self._context.close()
        await self._browser.close()
        await self._playwright.stop()


async def run(invoices, credentials):
    """Log in once, fill every invoice, then wait before closing the browser."""
    session = InvoiceSession()
    await session.start(credentials, on_event=lambda e: print(e.get("log", "")))
    results = await session.fill(invoices, on_event=print_event)

    # Keep the filled tabs on screen so they can be reviewed before the
    # window closes.
    input("Press Enter to close the browser...")
    await session.close()
    return results


def main() -> None:
    """Read the Excel file and fill every invoice, one tab per client."""
    load_dotenv()
    path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_EXCEL

    invoices = read_invoices(path)
    print(f"{len(invoices)} invoice(s) read from {path}")

    credentials = get_credentials()
    results = asyncio.run(run(invoices, credentials))
    done = results.count(True)
    skipped = results.count(False)
    failed = results.count(None)

    mode = "submitted" if SUBMIT else "dry-run"
    print(f"End ({mode}): {done} filled, {skipped} skipped, {failed} failed")
    if failed:
        sys.exit(1)


if __name__ == "__main__":
    main()
